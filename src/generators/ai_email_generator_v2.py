import os
import json
import logging
from typing import Dict, Any, Optional, List

# Importaciones opcionales para Excel
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logging.warning("Pandas no está disponible. Funcionalidad de Excel limitada.")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("OpenPyXL no está disponible. Funcionalidad de Excel limitada.")

# Importaciones opcionales para AI
try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False
    logging.warning("OpenAI no está disponible.")

try:
    import anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False
    logging.warning("Anthropic no está disponible.")

logger = logging.getLogger(__name__)

class AIEmailGeneratorV2:
    def __init__(self, openai_api_key: Optional[str] = None, anthropic_api_key: Optional[str] = None):
        """
        Inicializa el generador de emails AI con manejo robusto de dependencias
        
        Args:
            openai_api_key: Clave API de OpenAI (opcional)
            anthropic_api_key: Clave API de Anthropic (opcional)
        """
        self.openai_client = None
        self.anthropic_client = None
        
        # Configurar OpenAI si está disponible
        if OPENAI_AVAILABLE and openai_api_key:
            try:
                self.openai_client = openai.OpenAI(api_key=openai_api_key)
                logger.info("Cliente OpenAI configurado correctamente")
            except Exception as e:
                logger.error(f"Error configurando OpenAI: {e}")
        
        # Configurar Anthropic si está disponible
        if ANTHROPIC_AVAILABLE and anthropic_api_key:
            try:
                self.anthropic_client = anthropic.Anthropic(api_key=anthropic_api_key)
                logger.info("Cliente Anthropic configurado correctamente")
            except Exception as e:
                logger.error(f"Error configurando Anthropic: {e}")
    
    def load_excel_profile(self, excel_path: str) -> Dict[str, Any]:
        """
        Carga perfil de usuario desde archivo Excel con múltiples métodos de fallback
        
        Args:
            excel_path: Ruta al archivo Excel
            
        Returns:
            Dict con datos del perfil
        """
        # Intentar con pandas primero
        if PANDAS_AVAILABLE:
            try:
                return self._load_excel_pandas(excel_path)
            except Exception as e:
                logger.warning(f"Error con pandas, intentando método alternativo: {e}")
        
        # Fallback a openpyxl
        if OPENPYXL_AVAILABLE:
            try:
                return self._load_excel_openpyxl(excel_path)
            except Exception as e:
                logger.warning(f"Error con openpyxl: {e}")
        
        # Fallback final - archivo JSON si existe
        json_path = excel_path.replace('.xlsx', '.json').replace('.xls', '.json')
        if os.path.exists(json_path):
            try:
                return self._load_json_profile(json_path)
            except Exception as e:
                logger.error(f"Error cargando JSON: {e}")
        
        logger.error("No se pudo cargar el perfil con ningún método")
        return {}
    
    def _load_excel_pandas(self, excel_path: str) -> Dict[str, Any]:
        """Carga Excel usando pandas"""
        df = pd.read_excel(excel_path, sheet_name=0)
        
        if df.empty:
            logger.warning("El archivo Excel está vacío")
            return {}
        
        # Convertir primera fila a diccionario
        profile = df.iloc[0].to_dict()
        
        # Limpiar valores NaN
        profile = {k: v for k, v in profile.items() if pd.notna(v)}
        
        logger.info(f"Perfil cargado desde Excel con pandas: {list(profile.keys())}")
        return profile
    
    def _load_excel_openpyxl(self, excel_path: str) -> Dict[str, Any]:
        """Carga Excel usando openpyxl"""
        from openpyxl import load_workbook
        
        workbook = load_workbook(excel_path)
        sheet = workbook.active
        
        # Obtener headers de la primera fila
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value))
        
        # Obtener valores de la segunda fila
        values = []
        if sheet.max_row >= 2:
            for cell in sheet[2]:
                values.append(cell.value if cell.value is not None else "")
        
        # Crear diccionario
        profile = dict(zip(headers, values))
        
        logger.info(f"Perfil cargado desde Excel con openpyxl: {list(profile.keys())}")
        return profile
    
    def _load_json_profile(self, json_path: str) -> Dict[str, Any]:
        """Carga perfil desde JSON"""
        with open(json_path, 'r', encoding='utf-8') as f:
            profile = json.load(f)
        
        logger.info(f"Perfil cargado desde JSON: {list(profile.keys())}")
        return profile
    
    def save_profile_as_json(self, profile: Dict[str, Any], output_path: str) -> bool:
        """Guarda perfil como JSON para usar como fallback"""
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(profile, f, ensure_ascii=False, indent=2)
            
            logger.info(f"Perfil guardado como JSON en: {output_path}")
            return True
        except Exception as e:
            logger.error(f"Error guardando JSON: {e}")
            return False
    
    def generate_email(self, 
                      job_data: Dict[str, Any], 
                      user_data: Dict[str, Any],
                      excel_profile: Optional[Dict[str, Any]] = None,
                      template: Optional[str] = None) -> str:
        """
        Genera email personalizado usando AI con fallbacks robustos
        
        Args:
            job_data: Datos del trabajo
            user_data: Datos del usuario
            excel_profile: Perfil desde Excel (opcional)
            template: Template base (opcional)
            
        Returns:
            Email generado
        """
        # Combinar todos los datos de perfil
        profile = {**user_data}
        if excel_profile:
            profile.update(excel_profile)
        
        # Usar AI si está disponible
        try:
            if self.openai_client:
                return self._generate_with_openai(job_data, profile, template)
            elif self.anthropic_client:
                return self._generate_with_anthropic(job_data, profile, template)
        except Exception as e:
            logger.error(f"Error con AI: {e}")
        
        # Fallback a generación básica por template
        return self._generate_basic_email(job_data, profile, template)
    
    def _generate_with_openai(self, job_data: Dict[str, Any], profile: Dict[str, Any], template: Optional[str]) -> str:
        """Genera email usando OpenAI"""
        prompt = self._create_prompt(job_data, profile, template)
        
        response = self.openai_client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "Eres un experto en escribir cartas de presentación profesionales para trabajos en educación."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1000,
            temperature=0.7
        )
        
        return response.choices[0].message.content.strip()
    
    def _generate_with_anthropic(self, job_data: Dict[str, Any], profile: Dict[str, Any], template: Optional[str]) -> str:
        """Genera email usando Anthropic Claude"""
        prompt = self._create_prompt(job_data, profile, template)
        
        response = self.anthropic_client.messages.create(
            model="claude-3-sonnet-20240229",
            max_tokens=1000,
            messages=[
                {"role": "user", "content": prompt}
            ]
        )
        
        return response.content[0].text.strip()
    
    def _create_prompt(self, job_data: Dict[str, Any], profile: Dict[str, Any], template: Optional[str]) -> str:
        """Crea prompt para AI"""
        prompt = f"""
Genera una carta de presentación profesional en español para un trabajo en educación.

DATOS DEL TRABAJO:
- Título: {job_data.get('title', 'N/A')}
- Descripción: {job_data.get('description', 'N/A')}
- Empresa: {job_data.get('company', 'N/A')}
- Ubicación: {job_data.get('location', 'N/A')}

DATOS DEL CANDIDATO:
"""
        
        for key, value in profile.items():
            if value:
                prompt += f"- {key}: {value}\n"
        
        prompt += """
INSTRUCCIONES:
1. Crea una carta profesional y personalizada
2. Destaca la experiencia relevante del candidato
3. Conecta las habilidades del candidato con los requisitos del trabajo
4. Mantén un tono profesional pero cálido
5. Incluye un cierre apropiado
6. Máximo 300 palabras
7. No incluyas [Tu nombre] o placeholders, usa datos reales

FORMATO:
Asunto: [Asunto apropiado]

[Cuerpo del email]

Atentamente,
[Nombre del candidato]
"""
        
        return prompt
    
    def _generate_basic_email(self, job_data: Dict[str, Any], profile: Dict[str, Any], template: Optional[str]) -> str:
        """Genera email básico usando template"""
        try:
            if not template:
                template = self._get_default_template()
            
            # Reemplazar variables en el template de forma segura
            safe_profile = {k: str(v) if v else 'N/A' for k, v in profile.items()}
            safe_job = {k: str(v) if v else 'N/A' for k, v in job_data.items()}
            
            email_content = template.format(
                name=safe_profile.get('name', 'N/A'),
                email=safe_profile.get('email', 'N/A'),
                job_title=safe_job.get('title', 'Posición disponible'),
                company=safe_job.get('company', 'su organización'),
                location=safe_job.get('location', ''),
                experience=safe_profile.get('experience', 'mi experiencia profesional'),
                education=safe_profile.get('education', 'mi formación académica'),
                skills=safe_profile.get('skills', 'mis habilidades'),
                phone=safe_profile.get('phone', '')
            )
            
            return email_content
            
        except Exception as e:
            logger.error(f"Error generando email básico: {e}")
            return self._get_fallback_email(job_data, profile)
    
    def _get_default_template(self) -> str:
        """Template por defecto"""
        return """Estimado/a responsable de selección,

Me dirijo a ustedes para expresar mi interés en la posición de {job_title} en {company}.

Con {experience} y {education}, considero que mi perfil se ajusta perfectamente a los requisitos de esta posición. Mis {skills} me permiten contribuir efectivamente al equipo educativo.

Adjunto mi CV para su consideración y quedo a disposición para una entrevista.

Atentamente,
{name}
{email}
{phone}"""
    
    def _get_fallback_email(self, job_data: Dict[str, Any], profile: Dict[str, Any]) -> str:
        """Email de emergencia"""
        return f"""Estimado/a responsable de selección,

Me dirijo a ustedes para expresar mi interés en la posición de {job_data.get('title', 'educación')} en {job_data.get('company', 'su institución')}.

Adjunto mi CV para su consideración.

Atentamente,
{profile.get('name', 'Candidato')}
{profile.get('email', '')}"""
    
    def create_excel_template(self, output_path: str) -> bool:
        """
        Crea un template Excel de ejemplo con múltiples métodos
        
        Args:
            output_path: Ruta donde guardar el template
            
        Returns:
            True si se creó exitosamente
        """
        # Intentar con pandas primero
        if PANDAS_AVAILABLE:
            try:
                return self._create_excel_template_pandas(output_path)
            except Exception as e:
                logger.warning(f"Error creando template con pandas: {e}")
        
        # Fallback a openpyxl
        if OPENPYXL_AVAILABLE:
            try:
                return self._create_excel_template_openpyxl(output_path)
            except Exception as e:
                logger.warning(f"Error creando template con openpyxl: {e}")
        
        # Fallback final - crear JSON
        json_path = output_path.replace('.xlsx', '.json').replace('.xls', '.json')
        return self._create_json_template(json_path)
    
    def _create_excel_template_pandas(self, output_path: str) -> bool:
        """Crea template usando pandas"""
        template_data = {
            'name': ['Juan Pérez'],
            'email': ['juan.perez@email.com'],
            'phone': ['123-456-789'],
            'experience': ['5 años como profesor de matemáticas'],
            'education': ['Licenciatura en Matemáticas, Universidad XYZ'],
            'skills': ['Python, metodologías activas, manejo de aulas virtuales'],
            'specialization': ['Educación Secundaria'],
            'languages': ['Español (nativo), Inglés (avanzado)'],
            'certifications': ['Certificación en TIC educativas'],
            'motivation': ['Pasión por la enseñanza y el desarrollo estudiantil']
        }
        
        df = pd.DataFrame(template_data)
        df.to_excel(output_path, index=False)
        
        logger.info(f"Template Excel creado con pandas en: {output_path}")
        return True
    
    def _create_excel_template_openpyxl(self, output_path: str) -> bool:
        """Crea template usando openpyxl"""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Perfil Usuario"
        
        # Headers
        headers = ['name', 'email', 'phone', 'experience', 'education', 'skills', 
                  'specialization', 'languages', 'certifications', 'motivation']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Ejemplo de datos
        example_data = [
            'Juan Pérez', 'juan.perez@email.com', '123-456-789',
            '5 años como profesor de matemáticas',
            'Licenciatura en Matemáticas, Universidad XYZ',
            'Python, metodologías activas, manejo de aulas virtuales',
            'Educación Secundaria', 'Español (nativo), Inglés (avanzado)',
            'Certificación en TIC educativas',
            'Pasión por la enseñanza y el desarrollo estudiantil'
        ]
        
        for col, data in enumerate(example_data, 1):
            ws.cell(row=2, column=col, value=data)
        
        wb.save(output_path)
        
        logger.info(f"Template Excel creado con openpyxl en: {output_path}")
        return True
    
    def _create_json_template(self, output_path: str) -> bool:
        """Crea template como JSON"""
        template_data = {
            'name': 'Juan Pérez',
            'email': 'juan.perez@email.com',
            'phone': '123-456-789',
            'experience': '5 años como profesor de matemáticas',
            'education': 'Licenciatura en Matemáticas, Universidad XYZ',
            'skills': 'Python, metodologías activas, manejo de aulas virtuales',
            'specialization': 'Educación Secundaria',
            'languages': 'Español (nativo), Inglés (avanzado)',
            'certifications': 'Certificación en TIC educativas',
            'motivation': 'Pasión por la enseñanza y el desarrollo estudiantil'
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(template_data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"Template JSON creado en: {output_path}")
        return True
    
    def get_available_features(self) -> Dict[str, bool]:
        """Retorna qué características están disponibles"""
        return {
            'pandas': PANDAS_AVAILABLE,
            'openpyxl': OPENPYXL_AVAILABLE,
            'openai': OPENAI_AVAILABLE and self.openai_client is not None,
            'anthropic': ANTHROPIC_AVAILABLE and self.anthropic_client is not None,
            'excel_support': PANDAS_AVAILABLE or OPENPYXL_AVAILABLE,
            'ai_generation': (OPENAI_AVAILABLE and self.openai_client is not None) or 
                           (ANTHROPIC_AVAILABLE and self.anthropic_client is not None)
        }
